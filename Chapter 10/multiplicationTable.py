#! python3

import openpyxl, sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font

#Create a workbook
wb = openpyxl.Workbook()
sheet = wb.active

#Make named style for bold data
makingBold = NamedStyle(name='makingBold')
makingBold.font = Font(bold=True)

#Making entry for rows
row = 1
for column in range(1,int(sys.argv[1])+2):
	columnLetter = get_column_letter(column)
	if columnLetter+str(row) == 'A1':
		continue
	else:
		sheet[columnLetter+str(row)] = column-1
		sheet[columnLetter+str(row)].style = makingBold

#Making entry for columns
column=1
columnLetter = get_column_letter(column)
for row in range(1,int(sys.argv[1])+2):
	if columnLetter+str(row) == 'A1':
		continue
	else:
		sheet[columnLetter+str(row)] = row-1
		sheet[columnLetter+str(row)].style = makingBold
		
#Making rest of the entry		
rowLetter = get_column_letter(1)
for row in range(2,int(sys.argv[1])+2):
	for column in range(2,int(sys.argv[1])+2):
		columnLetter = get_column_letter(column)
		sheet[columnLetter+str(row)] = sheet[columnLetter+str(1)].value * sheet[rowLetter+str(row)].value

#Save the file
wb.save('D:\\Scripts\\Multiplication_Table.xlsx')
