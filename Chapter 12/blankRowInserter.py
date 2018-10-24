#! python3

import openpyxl, sys, os
from openpyxl.utils import get_column_letter

#Passed parameters
N = int(sys.argv[1]) #Start position of blank rows
M = int(sys.argv[2]) #Number of blank rows to insert
filename = sys.argv[3]

#Loading the sheet for reading
readingBook = openpyxl.load_workbook('D:\\Scripts\\produceSalesCopy.xlsx')
readingSheet = readingBook.active

#Creating the sheet for writing
writingBook = openpyxl.Workbook()
writingSheet = writingBook.active

#Writing first N lines
for row in range(1,N):
	for column in range(1,readingSheet.max_column+1):
		columnLetter = get_column_letter(column)
		writingSheet[columnLetter+str(row)] = readingSheet[columnLetter+str(row)].value
		
#Writing next batch of lines after giving space of M number of lines
for row in range(N,readingSheet.max_row+1):
	for column in range(1,readingSheet.max_column+1):
		columnLetter = get_column_letter(column)
		writingSheet[columnLetter+str(row+M)] = readingSheet[columnLetter+str(row)].value

#Saving the workbook
writingBook.save('D:\\Scripts\\'+filename)
