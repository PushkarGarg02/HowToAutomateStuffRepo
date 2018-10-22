#! python3

import openpyxl

print('Updating the excel with correct values...')

PRICE_UPDATES = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon':1.27}

wb = openpyxl.load_workbook('D:\\Scripts\\produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

for row in range(2,sheet.max_row):
	produceName = sheet.cell(row=row,column=1).value
	if produceName in PRICE_UPDATES:
		sheet.cell(row=row,column=2).value = PRICE_UPDATES[produceName]
		#sheet.cell(row=row,column=4).value = float(PRICE_UPDATES[produceName]) * float(sheet.cell(row=row,column=4).value)
		
wb.save('D:\\Scripts\\updatedProduceSalesPushkar.xlsx')
