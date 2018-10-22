#! python3

import openpyxl
from openpyxl.styles import Font, NamedStyle
highlight = NamedStyle(name='highlight')
highlight.font = Font(size=24,italic=True, name='Times New Roman')

#Open workbook and assign a style
#wb = openpyxl.load_workbook('D:\\Scripts\\styled.xlsx')
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'].style = highlight
sheet['A1'] = 'Hello World Pushkar!'
wb.save('D:\\Scripts\\styled123.xlsx')

#Open workbook and assign the formula
wb1 = openpyxl.Workbook()
sheet1 = wb1.active
sheet1['A1'] = 200
sheet1['A2'] = 300
sheet1['A3'] = '=SUM(A1:A2)'
wb1.save('D:\\Scripts\\writeFormula.xlsx')

#Print the formula in the cell without parameter data_only=True
wbFormulas = openpyxl.load_workbook('D:\\Scripts\\writeFormula.xlsx')
sheet = wbFormulas.active
print(sheet['A3'].value)

#Print the actual data in the cell
wbDataOnly = openpyxl.load_workbook('D:\\Scripts\\writeFormula.xlsx', data_only=True)
sheet = wbDataOnly.active

#Set the height of rows and columns
sheet.row_dimensions[1].height=70
sheet.column_dimensions['B'].width = 20

#Merge the cells
sheet.merge_cells('A4:D5')
sheet['A4'] = 'Here is the merging'

#Freeze the panes
sheet.freeze_panes = 'A2'

#Un-freeze panes
sheet.freeze_panes = 'A1'
sheet.freeze_panes = None

print(sheet['A3'].value)
wbDataOnly.save('D:\\Scripts\\updatedWriteFormula.xlsx')
